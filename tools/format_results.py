#!/usr/bin/env python3
"""Gerar XLSX formatado a partir de um CSV de resultados.

Uso:
    python tools\format_results.py outputs\resultados-1.csv

Gera `outputs/resultados-1.xlsx` com cabeçalho em negrito, filtro, congelamento de painel
e coloração condicional simples baseada na coluna `classificacao`.
"""
import argparse
import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter


CLASS_COLOR = {
    'OK': 'C6EFCE',
    'NAO_ENCONTRADO': 'FFC7CE',
    'VERIFICAR_NUMNOTA': 'FFEB9C',
    'VERIFICAR_VALOR': 'FFEB9C',
    'HOTEL': 'F4CCCC',
    'SEM_VALOR': 'FCE4D6',
}
ERROR_COLOR = 'F8696B'  # red-ish for combined/serious errors


def read_csv(path):
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f, skipinitialspace=True)
        rows = [r for r in reader if any((v and str(v).strip()) for v in r.values())]
        return reader.fieldnames or [], rows


def make_xlsx(csv_path, out_path=None):
    fieldnames, rows = read_csv(csv_path)
    # normalize/flag value differences: if original 'Valor' differs from 'matched_valor', mark as VERIFICAR_VALOR
    def parse_currency(s):
        if s is None:
            return None
        s = str(s).strip()
        if s == '':
            return None
        s = s.replace('"', '').replace('R$', '').replace(' ', '')
        # handle formats like 47,00 or 47.0
        # first try to replace comma decimal
        try:
            if ',' in s and '.' in s:
                # assume thousand separator is '.' and decimal is ','
                s2 = s.replace('.', '').replace(',', '.')
            else:
                s2 = s.replace(',', '.')
            return float(s2)
        except Exception:
            try:
                return float(s)
            except Exception:
                return None

    for row in rows:
        orig_s = row.get('Valor') or row.get('valor') or ''
        matched_s = row.get('matched_valor') or row.get('matched_valor') or ''
        orig = parse_currency(orig_s)
        matched = parse_currency(matched_s)
        if orig is not None and matched is not None:
            # flag verification only if difference exceeds 0.70 (70 centavos)
            try:
                if abs(orig - matched) > 0.70:
                    row['classificacao'] = 'VERIFICAR_VALOR'
                    prev = row.get('observacao') or ''
                    if 'VERIFICAR_VALOR' not in prev:
                        row['observacao'] = (prev + ' | VERIFICAR_VALOR').strip(' |')
            except Exception:
                pass
    if not out_path:
        out_path = os.path.splitext(csv_path)[0] + '.xlsx'
    os.makedirs(os.path.dirname(out_path) or '.', exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = 'resultados'

    # header
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    for c, h in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # write rows
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, h in enumerate(fieldnames, start=1):
            val = row.get(h, '')
            ws.cell(row=r_idx, column=c_idx, value=val)

    # column widths (basic auto width)
    for i, h in enumerate(fieldnames, start=1):
        col = get_column_letter(i)
        max_len = max((len(str(ws[f"{col}{r}"].value or '')) for r in range(1, ws.max_row + 1)), default=0)
        ws.column_dimensions[col].width = min( max(10, max_len + 2), 50)

    # freeze header and autofilter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(fieldnames))}{ws.max_row}'

    # apply classification colors if column exists
    if 'classificacao' in fieldnames:
        col_idx = fieldnames.index('classificacao') + 1
        # find indices for potential columns to inspect
        try:
            valor_col_idx = fieldnames.index('matched_valor') + 1
        except ValueError:
            valor_col_idx = None
        try:
            orig_val_col_idx = fieldnames.index('Valor') + 1
        except ValueError:
            orig_val_col_idx = None
        obs_col_idx = None
        try:
            obs_col_idx = fieldnames.index('observacao') + 1
        except ValueError:
            obs_col_idx = None

        def _parse_numcell(v):
            if v is None:
                return None
            s = str(v).strip().replace('"', '').replace('R$', '').replace(' ', '')
            if s == '':
                return None
            s = s.replace('.', '') if s.count('.') > 1 else s
            s = s.replace(',', '.')
            try:
                return float(s)
            except Exception:
                try:
                    return float(s)
                except Exception:
                    return None

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            key = (cell.value or '').strip()
            obs = (ws.cell(row=r, column=obs_col_idx).value or '') if obs_col_idx else ''

            # compute numeric diff if possible
            orig_val = None
            matched_val = None
            if orig_val_col_idx:
                orig_val = _parse_numcell(ws.cell(row=r, column=orig_val_col_idx).value)
            if valor_col_idx:
                matched_val = _parse_numcell(ws.cell(row=r, column=valor_col_idx).value)

            color = None
            # if both VERIFICAR_NUMNOTA and VERIFICAR_VALOR present -> serious error (red)
            combined_flag = ('VERIFICAR_NUMNOTA' in key) or ('VERIFICAR_NUMNOTA' in (obs or ''))
            value_flag = ('VERIFICAR_VALOR' in key) or ('VERIFICAR_VALOR' in (obs or ''))
            if combined_flag and value_flag:
                color = ERROR_COLOR
            else:
                # if numeric difference is very large (e.g. > 5.0), mark as error too
                try:
                    if orig_val is not None and matched_val is not None and abs(orig_val - matched_val) > 5.0:
                        color = ERROR_COLOR
                except Exception:
                    pass

            # fallback to normal coloring
            if color is None:
                # Treat ALMOCO and JANTA as OK for coloring purposes (both get green)
                if key in ('ALMOCO', 'JANTA'):
                    color = CLASS_COLOR.get('OK')
                else:
                    color = CLASS_COLOR.get(key)

            if color:
                for c in range(1, len(fieldnames) + 1):
                    ws.cell(row=r, column=c).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

    wb.save(out_path)
    return out_path


def main():
    p = argparse.ArgumentParser()
    p.add_argument('csv', nargs='?', default=os.path.join('outputs', 'resultados-1.csv'))
    p.add_argument('-o', '--out', help='Arquivo xlsx de saída')
    args = p.parse_args()

    if not os.path.exists(args.csv):
        print('CSV not found:', args.csv)
        return
    out = make_xlsx(args.csv, out_path=args.out)
    print('Wrote', out)


if __name__ == '__main__':
    main()
